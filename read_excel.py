import pandas as pd
import openpyxl as op
from openpyxl.utils import get_column_letter


#.head(n)
#.tale(n)
#.columns -> headers and type
#.dtypes -> types of columns
#['name'] -> select by header
#.at[row, header] -> select one item
#.loc[row1 : row2, header1 : header2] -> select a matrix of items, [row1 ::step]
#df[sheetname] -> select sheet
class read_data:
    def __init__(self, path : str):
        self.path = path
    
    def read_xslx(self, sheetname : int|str = 0, header : int|list[int]|None = 0, names: list | None = None, usecol : list[any] = None, nrows : int = None):
        self.dataframe = pd.read_excel(self.path, sheet_name=sheetname, header=header, names=names, usecols=usecol ,nrows=nrows)
    
    def read_json(self):
        self.dataframe = pd.read_json(self.path)
    
    def save_json(self, file):
        self.dataframe.to_json(file)
    
    def read_csv(self):
        self.dataframe = pd.read_csv(self.path)
    
    def save_csv(self, file):
        self.dataframe.to_csv(file)
    
    def read_clipboard(self):
        self.dataframe = pd.read_clipboard()
    
    def save_clipboard(self):
        self.dataframe.to_clipboard()
    
    def read_html(self):
        self.dataframe = pd.read_html(self.path)

    def save_html(self, file):
        self.dataframe.to_html(file)
        
    def get_headers(self):
        return self.dataframe.columns

    def get_column(self, header : str):
        return self.dataframe[header]
    
    def get_block(self, row : int, header : str): #one var
        return self.dataframe.at[row, header]
    
    def get_list(self, rowrange, headerrange): #matrix of vars
        return self.dataframe.loc[rowrange[0] : rowrange[1], headerrange[0] : headerrange[1]]

    
    def convert_to_ttk_table(self)->list[list]: #add read html option
        data = [] #[[headers],[data],[data]...]
        data.append(list(self.dataframe.columns))
        for i in range(len(self.dataframe)):
            row = []
            for key in self.dataframe:
                row.append(self.dataframe.get(key)[i])
                
            data.append(row)
        return data
        

class modify_data:
    MAX_ROW = [1]
    MAX_COLUMN = [1]
    active_sheet = 0
    sheet_count = 0
    def __init__(self, name : str, mode : bool = True):
        self.workBook = op.Workbook()
        self.workBook.remove(self.workBook['Sheet'])
        self.name = name
        self.saveMode = mode
    
    def create_Sheet(self, title : str|None = ''):
        #create a new sheet
        self.workSheet = self.workBook.create_sheet(title=title)
        self.sheet_count += 1
        self.MAX_ROW.append(1)
        self.MAX_COLUMN.append(1)
    
    def remove_Sheet(self, sheetindex):
        self.workBook.remove(self.workBook.worksheets[sheetindex])
        self.sheet_count -= 1

    def add_row(self, data : list):
        #add a row
        self.workBook[self.workBook.sheetnames[self.active_sheet]].append(data)
        self.MAX_ROW[self.active_sheet] += 1
    
    def add_col(self, data : list):
        try:
            for row in range(self.MAX_ROW[self.active_sheet], len(data) + self.MAX_ROW[self.active_sheet]):
                self.workBook[self.workBook.sheetnames[self.active_sheet]][get_column_letter(self.MAX_COLUMN[self.active_sheet]) + str(row)] = data[row - self.MAX_ROW[self.active_sheet]]
            self.MAX_COLUMN[self.active_sheet] += 1
        except ValueError:
            print('value error')

    def loadExcel(self,path : str):
        try:
            self.workBook = op.load_workbook(path)
            for i in range(len(self.workBook.worksheets)):

                self.MAX_ROW[i] = self.workBook[self.workBook.worksheets[i]].max_row
                self.MAX_COLUMN[i] = self.workBook[self.workBook.worksheets[i]].max_column

        except FileNotFoundError:
            print('process failed.')
        
    def Loaddict(self, data : dict, sheetName : str = 'Sheet1', mode : str = 'V' ):
        self.workSheet.title = sheetName
        self.data = data
        if mode == 'H':
            name = ''
            self.header = [name] + [header for header in data[list(data)[0]].keys()]
            self.workSheet.append(self.header)
            for element in data:
                self.workSheet.append([element] + [block for block in data[element].values()])
            
            self.MAX_ROW[self.active_sheet] = self.workSheet.max_row
            self.MAX_COLUMN[self.active_sheet] = self.workSheet.max_column
        
        elif mode == 'V':
            self.header =  [element for element in data]
            self.add_row(self.header)
            for header in self.header:
                self.add_col([value for value in data[header].values()])
            self.MAX_ROW[self.active_sheet] = self.worksheet.max_row
            self.MAX_COLUMN[self.active_sheet] = self.worksheet.max_column

    def LoadList(self, data : list[list], sheetName : str = 'Sheet1', file_type : str = 'xlsx'):
        self.data = data
        for row in data:
            self.add_row(row)
        self.MAX_ROW[self.active_sheet] = self.workSheet.max_row
        self.MAX_COLUMN[self.active_sheet] = self.workSheet.max_column
        self.save(file_type)

    def save(self, file_type):
        self.workBook.save(self.name + f'.{file_type}')

class save_excel:
    def __init__(self, **kargs) -> None:
        workBook = op.Workbook()
        workBook.remove(workBook['Sheet'])
        workSheet = workBook.create_sheet('saved')
        data : list[list] = kargs.get('data')
        name : str = kargs.get('name')
        for row in data:
            workBook['saved'].append(row)
        workBook.save(name + '.xlsx')

if __name__ == '__main__':
    #FSI-2023-DOWNLOAD
    df = read_data('mouseInfoScreenshot.png')
    df.read_xslx()
    print(len(df.dataframe))
    print(df.dataframe)

   
    # df.clear_df()
    # print(df.dataframe)
    # print(df.dataframe)
    # print(df.dataframe['Sheet1']['z'][0])
